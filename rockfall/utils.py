"""
工具层 — 公共函数
==================
全项目共享的底层工具函数。
"""

import io
import json
from datetime import datetime

import numpy as np


def box_iou_batch(boxes_a: np.ndarray, boxes_b: np.ndarray) -> np.ndarray:
    """
    批量计算两组边界框的 IoU 矩阵。

    boxes_a: (N, 4) [x1, y1, x2, y2]
    boxes_b: (M, 4) [x1, y1, x2, y2]
    返回:   (N, M) IoU 矩阵
    """
    x1 = np.maximum(boxes_a[:, None, 0], boxes_b[None, :, 0])
    y1 = np.maximum(boxes_a[:, None, 1], boxes_b[None, :, 1])
    x2 = np.minimum(boxes_a[:, None, 2], boxes_b[None, :, 2])
    y2 = np.minimum(boxes_a[:, None, 3], boxes_b[None, :, 3])
    inter = np.maximum(0, x2 - x1) * np.maximum(0, y2 - y1)

    area_a = (boxes_a[:, 2] - boxes_a[:, 0]) * (boxes_a[:, 3] - boxes_a[:, 1])
    area_b = (boxes_b[:, 2] - boxes_b[:, 0]) * (boxes_b[:, 3] - boxes_b[:, 1])
    union = area_a[:, None] + area_b[None, :] - inter

    return inter / np.maximum(union, 1e-6)


# ============================================================
# Excel 归档导出 (应急管理部门合规要求)
# ============================================================

# 四级预警中文标签
_LEVEL_LABELS_CN = {
    "red":    "Ⅰ级·特别严重",
    "orange": "Ⅱ级·严重",
    "yellow": "Ⅲ级·较重",
    "blue":   "Ⅳ级·一般",
}

# Excel 列定义: (字段名, 列宽, 中文标题)
_EXPORT_COLUMNS = [
    ("id",                  8,  "序号"),
    ("time",               20,  "报警时间"),
    ("monitoring_location",18,  "监测点位"),
    ("alert_level_cn",     16,  "预警等级"),
    ("count",              10,  "落石数量"),
    ("max_confidence",     12,  "最高置信度"),
    ("rock_diameter_cm",   12,  "落石直径(cm)"),
    ("class_summary",      18,  "检测类别"),
    ("push_status_cn",     14,  "推送状态"),
    ("saved_frame",        45,  "截图路径"),
    ("created_at",         20,  "入库时间"),
]

_PUSH_STATUS_CN = {
    "recorded": "仅记录",
    "popup":    "已弹窗",
    "sent":     "已推送",
    "pending":  "待推送",
    "failed":   "推送失败",
}


def export_alerts_to_excel(alerts: list[dict], sheet_title: str = "落石预警记录") -> bytes:
    """
    将预警记录列表导出为格式化 Excel (.xlsx) 字节, 符合应急管理部门归档要求。

    参数:
        alerts:      预警记录 dict 列表 (来自 alert_store)
        sheet_title: 工作表标题

    返回:
        .xlsx 文件字节

    Excel 格式:
      - 标题行冻结 + 自动筛选
      - 四级预警等级颜色标记 (红/橙/黄/蓝)
      - 列宽自适应
      - 表头加粗居中
      - 边框线
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError(
            "导出 Excel 需要 openpyxl 库。请执行: pip install openpyxl"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "预警记录"

    # ---- 标题行 ----
    title_row = 1
    ws.merge_cells(start_row=title_row, start_column=1,
                   end_row=title_row, end_column=len(_EXPORT_COLUMNS))
    title_cell = ws.cell(row=title_row, column=1, value=sheet_title)
    title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[title_row].height = 32

    # 副标题: 导出时间 + 记录条数
    sub_row = 2
    ws.merge_cells(start_row=sub_row, start_column=1,
                   end_row=sub_row, end_column=len(_EXPORT_COLUMNS))
    export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sub_cell = ws.cell(row=sub_row, column=1,
                       value=f"导出时间: {export_time}    共 {len(alerts)} 条记录")
    sub_cell.font = Font(name="微软雅黑", size=9, color="666666")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sub_row].height = 22

    # ---- 表头 (第3行) ----
    header_row = 3
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    for col_idx, (_, width, title) in enumerate(_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 22

    # ---- 数据行 ----
    level_fills = {
        "red":    PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # 浅红
        "orange": PatternFill(start_color="FFD8B0", end_color="FFD8B0", fill_type="solid"),  # 浅橙
        "yellow": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # 浅黄
        "blue":   PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # 浅蓝
    }
    level_fonts = {
        "red":    Font(name="微软雅黑", size=10, bold=True, color="9C0006"),
        "orange": Font(name="微软雅黑", size=10, bold=True, color="BF5B00"),
        "yellow": Font(name="微软雅黑", size=10, bold=True, color="9C6500"),
        "blue":   Font(name="微软雅黑", size=10, color="1F4E79"),
    }
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    for row_offset, alert in enumerate(alerts):
        row = header_row + 1 + row_offset
        level = alert.get("alert_level", "")
        ws.row_dimensions[row].height = 20

        for col_idx, (field, _, _) in enumerate(_EXPORT_COLUMNS, start=1):
            value = _format_cell_value(field, alert)
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

            # 预警等级列特殊着色
            if field == "alert_level_cn" and level in level_fills:
                cell.fill = level_fills[level]
                cell.font = level_fonts.get(level, data_font)
                cell.alignment = center_align
            elif field in ("id", "count", "max_confidence", "rock_diameter_cm"):
                cell.alignment = center_align
            elif field == "push_status_cn":
                cell.alignment = center_align

    # ---- 冻结表头 + 自动筛选 ----
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(_EXPORT_COLUMNS))}{header_row + len(alerts)}"

    # ---- 底部汇总行 ----
    summary_row = header_row + len(alerts) + 1
    ws.merge_cells(start_row=summary_row, start_column=1,
                   end_row=summary_row, end_column=len(_EXPORT_COLUMNS))
    level_counts = {}
    for a in alerts:
        lv = a.get("alert_level", "")
        level_counts[lv] = level_counts.get(lv, 0) + 1
    summary_parts = "  |  ".join(
        f"{_LEVEL_LABELS_CN.get(lv, lv)}: {level_counts[lv]}条"
        for lv in ["red", "orange", "yellow", "blue"]
        if lv in level_counts
    )
    summary_cell = ws.cell(row=summary_row, column=1,
                           value=f"汇总: {summary_parts}" if summary_parts else "无预警记录")
    summary_cell.font = Font(name="微软雅黑", size=9, italic=True, color="666666")
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")

    # ---- 写入 BytesIO ----
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _format_cell_value(field: str, alert: dict) -> str:
    """将 DB 原始值格式化为 Excel 可读的中文/数值"""
    level = alert.get("alert_level", "")

    if field == "alert_level_cn":
        return _LEVEL_LABELS_CN.get(level, level.upper() if level else "-")

    if field == "push_status_cn":
        ps = alert.get("push_status", "")
        return _PUSH_STATUS_CN.get(ps, ps or "-")

    if field == "rock_diameter_cm":
        val = alert.get("rock_diameter_cm", 0)
        try:
            v = float(val) if val else 0
            return f"{v:.1f}" if v > 0 else "-"
        except (ValueError, TypeError):
            return str(val) if val else "-"

    if field == "max_confidence":
        val = alert.get("max_confidence", 0)
        try:
            return round(float(val), 4)
        except (ValueError, TypeError):
            return val

    if field == "track_ids":
        val = alert.get("track_ids", "[]")
        if isinstance(val, str):
            try:
                parsed = json.loads(val)
                return ", ".join(str(t) for t in parsed) if parsed else "-"
            except (json.JSONDecodeError, TypeError):
                return val
        if isinstance(val, list):
            return ", ".join(str(t) for t in val) if val else "-"
        return str(val)

    val = alert.get(field, "")
    return val if val not in (None, "") else "-"
